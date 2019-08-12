import csv
import sys
import requests
import json
import datetime
import traceback
import os
import pandas
import time
import random
import lxml
import threading
import openpyxl
from lxml import html

alldata = []
search = ['a', 'b', 'c', 'd', 'e', 'f', 'g', 'h', 'i', 'j', 'k', 'l', 'm', 'n', 'o', 'p', 'q', 'r', 's', 't', 'u',
          'v', 'w', 'x', 'y', 'z', '1', '2', '3', '4', '5', '6', '7', '8', '9', '0']
stop_saving_requested = False
saving_running = False

def getinputfile(filename):
    wb = openpyxl.load_workbook('Source_eBay.xlsx')
    data_sheet = wb.worksheets[0]
    data = []
    for i in range(2, data_sheet.max_row):
        seller = []
        for b in range(1, 11):
            if data_sheet.cell(i, b).value is None:
                seller.append('')
            else:
                seller.append(data_sheet.cell(i, b).value)
        data.append(seller)
    return data

def isinalldata(link):
    for d in alldata:
        if link == d[0]:
            return True
    return False

def getsellers(link, index):
    if not isinalldata(link):
        print('New seller found, retrieving info')
        profpage= requests.get(link).text
        pphtml = html.fromstring(profpage)
        try:
            Seller_Name = pphtml.xpath('//h1[@class="str-billboard__title"]/text()')[0]
            seller_profile_url = pphtml.xpath('//div[@class="str-billboard__bsf"]/a/@href')[0]
        except IndexError:
            Seller_Name = link.split('/')[-1]
            seller_profile_url = 'Not Available'
            Business_name = Seller_Name
            First_name = 'Not Available'
            Last_name = 'Not Available'
            Address = 'Not Available'
            Phone = 'Not Available'
            Email = 'Not Available'
            Status = 'New_Deleted'
            alldata.append([link, Seller_Name, seller_profile_url, Business_name, First_name, Last_name, Address, Phone, Email, Status])
            return
        try:
            profpage = requests.get(seller_profile_url).text
            pphtml = html.fromstring(profpage)
        except:
            alldata.append([link, Seller_Name, seller_profile_url, 'Not Available', 'Not Available', 'Not Available', 'Not Available', 'Not Available', 'Not Available', 'New'])
            return
        if 'title="Business details"' in profpage:
            try:
                Business_name = pphtml.xpath('//div[@class="bsi_table"]/div[@class="bsi_row"]/span[@id="business_name"]/following-sibling::span/text()')[0]
            except IndexError:
                Business_name = Seller_Name
            try:
                First_name = pphtml.xpath('//div[@class="bsi_table"]/div[@class="bsi_row"]/span[@id="first_name"]/following-sibling::span/text()')[0]
            except IndexError:
                First_name = 'Not Available'
            try:
                Last_name = pphtml.xpath('//div[@class="bsi_table"]/div[@class="bsi_row"]/span[@id="last_name"]/following-sibling::span/text()')[0]
            except IndexError:
                Last_name = 'Not Available'
            try:
                Address_parts = pphtml.xpath('//span[@id="address"]/following-sibling::span/span/text()')
                Address = ', '.join(Address_parts).strip()
            except IndexError:
                Address = 'Not Available'
            try:
                Phone = pphtml.xpath('//div[@class="bsi_table"]/div[@class="bsi_row"]/span[@id="phone_number"]/following-sibling::span/text()')[0]
            except IndexError:
                Phone = 'Not Available'
            try:
                Email = pphtml.xpath('//div[@class="bsi_table"]/div[@class="bsi_row"]/span[@id="email"]/following-sibling::span/text()')[0]
            except IndexError:
                Email = 'Not Available'
            Status = 'New'
            alldata.append([link, Seller_Name, seller_profile_url, Business_name, First_name, Last_name, Address, Phone, Email, Status])
            return
        else:
            alldata.append([link, Seller_Name, seller_profile_url, 'Not Available', 'Not Available', 'Not Available', 'Not Available', 'Not Available', 'Not Available', 'New'])
            return
    p = requests.get(link).text
    phtml = html.fromstring(p)
    if 'This page does not exist' in p and 'http' in alldata[index][2]:
        print(link + ' Deleted')
        alldata[index][9] = 'Deleted'
        return
    if not alldata[index][2] in p and 'http' in alldata[index][2] and len(phtml.xpath('//div[@class="str-billboard__bsf"]/a/@href')) < 1:
        print(alldata[index][2] + ' Deleted')
        alldata[index][9] = 'Deleted'
        return

def getlinks(q):
    allresults = []
    i = 1
    while True:
        print('Getting results from page number ' + str(i) + ', searching "' + q + '"')
        page = requests.get('https://www.ebay.co.uk/sns?_pgn=' + str(i) + '&store_search=' + q).text
        results = html.fromstring(page).xpath('//li[@class="sns-item"]/div/a/@href')
        if (len(results) == 1 and results[0] == 'https://www.ebay.co.uk/str/' + q) or ("We couldn't find any shops with the name" in page):
            break
        else:
            allresults += results
        i+=1
    return allresults

def thread_working(index):
    try:
        s = search[index]
    except IndexError:
        print('Thread number ' + str(index) + ' has finished its work')
        return
    print('Getting results for thread number ' + str(index) + '\n')
    results = getlinks(s)
    print('Thread number ' + str(index) + ' got ' + str(len(results)) + ' results')
    for i, res in enumerate(results):
        print('Checking result number ' + str(i) + ' out of ' + str(len(results)) + ' in thread number ' + str(index))
        d = None
        for q, w in enumerate(alldata):
            if w[0] == res:
                d = q
        getsellers(res, d)
    print('Thread number ' + str(index) + ' has finished its work')

def results_saving():
    while True:
        time.sleep(5)
        wb = openpyxl.load_workbook('Source_eBay.xlsx')
        data_sheet = wb.worksheets[0]
        row_index = 2
        for seller in alldata:
            for column_index, element in enumerate(seller):
                data_sheet.cell(row_index, column_index+1).value = str(element)
            row_index += 1
        wb.save('Source_eBay.xlsx')
        if stop_saving_requested:
            saving_running = False
            return
        

alldata = getinputfile('Source_eBay.xlsx')
threads = []
for i, r in enumerate(search):
    t = threading.Thread(target=thread_working, args=(i,))
    threads.append(t)
    t.start()
s_thread = threading.Thread(target=results_saving)
s_thread.start()
saving_running = True
while True:
    if len(list(t for t in threads if t.is_alive())) == 0:
        print('All threads have finished its work')
        stop_saving_requested = True
        break
while True:
    if not saving_running:
        print('Script finished its work')
        break
